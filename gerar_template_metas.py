"""
Gera o template de metas para o dashboard AgentPerformanceScore.

Aba "metas" — colunas esperadas pelo dashboard:
  operacao | indicador | meta | dt_inicio | dt_fim

Regras de indicadores por tipo de operação:
  VENDA / RENOVAÇÃO / ASSESSORIA → Produtividade, Conversão, Monitoria
  ENDOSSO (VOZ E DIGITAL)        → Upselling, Cross Selling, Monitoria
  ENDOSSO FROTA                  → Upselling, Monitoria
  FROTA (VENDA E RENOVAÇÃO)      → Produtividade, Monitoria
  RETENÇÃO                       → Retenção, Cross Selling, Monitoria
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Mapeamento operacao_sgdot → indicadores ativos
# ─────────────────────────────────────────────
OPERACOES = {
    "Auto Venda Digital":     ["Produtividade", "Conversão",    "Monitoria"],
    "Auto Venda":             ["Produtividade", "Conversão",    "Monitoria"],
    "Auto Renovação Digital": ["Produtividade", "Conversão",    "Monitoria"],
    "Auto Renovação":         ["Produtividade", "Conversão",    "Monitoria"],
    "Auto Monitorada":        ["Produtividade", "Conversão",    "Monitoria"],
    "Auto Assessoria":        ["Produtividade", "Conversão",    "Monitoria"],
    "Auto Retenção":          ["Retenção",      "Cross Selling","Monitoria"],
    "Auto Pós-Venda":         ["Upselling",     "Cross Selling","Monitoria"],
    "Auto Endosso Digital":   ["Upselling",     "Cross Selling","Monitoria"],
    "Auto Endosso Frota":     ["Upselling",     "Monitoria"],
    "Auto Frota":             ["Produtividade", "Monitoria"],
}

# Meta diária padrão (referência inicial — ajustar conforme necessidade)
META_PADRAO = {
    "Produtividade": 10,
    "Conversão":     0.22,
    "Upselling":     4,
    "Cross Selling": 3,
    "Retenção":      4,
    "Monitoria":     90,
}

# ─────────────────────────────────────────────
# Estilos
# ─────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN = Alignment(horizontal="center", vertical="center")

OBS_FILL  = PatternFill("solid", fgColor="F0F4FA")
OBS_FONT  = Font(italic=True, color="4A5568", size=9)

ROW_FILLS = ["FFFFFF", "EEF3FB"]   # alternating

thin = Side(style="thin", color="CBD5E0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

NUM_FMT = {
    "Conversão":  "0.00%",
    "Monitoria":  "0.0",
}

def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_header(ws, headers):
    ws.row_dimensions[1].height = 28
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border = BORDER

# ─────────────────────────────────────────────
# Gerar workbook
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Aba principal: metas ──────────────────────
ws = wb.active
ws.title = "metas"
ws.freeze_panes = "A2"

headers = ["operacao", "indicador", "meta", "dt_inicio", "dt_fim"]
write_header(ws, headers)

col_width(ws, 1, 30)   # operacao
col_width(ws, 2, 18)   # indicador
col_width(ws, 3, 12)   # meta
col_width(ws, 4, 14)   # dt_inicio
col_width(ws, 5, 14)   # dt_fim

DATA_INICIO = "2026-01-01"
DATA_FIM    = ""          # vazio = sem limite

row = 2
fill_idx = 0
for operacao, indicadores in OPERACOES.items():
    for indicador in indicadores:
        meta = META_PADRAO.get(indicador, 0)
        fill = PatternFill("solid", fgColor=ROW_FILLS[fill_idx % 2])

        values = [operacao, indicador, meta, DATA_INICIO, DATA_FIM]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill   = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if c == 3 else "left")
        # Formato especial para conversão (%)
        if indicador in NUM_FMT:
            ws.cell(row=row, column=3).number_format = NUM_FMT[indicador]

        row += 1
        fill_idx += 1

# Linha curiga: wildcard para todas as operações
ws.cell(row=row, column=1, value="").fill = PatternFill("solid", fgColor=ROW_FILLS[fill_idx%2])
ws.cell(row=row, column=2, value="").fill = PatternFill("solid", fgColor=ROW_FILLS[fill_idx%2])

# Nota de rodapé (linha após os dados)
obs_row = row + 2
ws.merge_cells(f"A{obs_row}:E{obs_row}")
obs = ws.cell(row=obs_row, column=1,
              value=(
                  "INSTRUÇÕES: (1) 'operacao' = valor exato do campo operacao_sgdot  "
                  "(2) 'meta' = meta DIÁRIA — o dashboard multiplica pelo nº de dias trabalhados  "
                  "(3) 'dt_fim' pode ficar em branco (sem limite)  "
                  "(4) Para meta global (todos as operações) deixe 'operacao' em branco"
              ))
obs.font  = OBS_FONT
obs.fill  = OBS_FILL
obs.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[obs_row].height = 48

# ── Aba de referência ─────────────────────────
ref = wb.create_sheet("referencia")
ref.freeze_panes = "A2"
ref.sheet_view.showGridLines = True

ref_headers = ["operacao_sgdot", "tipo_operacao", "indicadores_ativos", "pesos"]
write_header(ref, ref_headers)
col_width(ref, 1, 26)
col_width(ref, 2, 32)
col_width(ref, 3, 38)
col_width(ref, 4, 40)

TIPOS = {
    "Auto Venda Digital":     ("VENDA / RENOVAÇÃO / ASSESSORIA", "Prod 60% | Conv 30% | Mon 10%"),
    "Auto Venda":             ("VENDA / RENOVAÇÃO / ASSESSORIA", "Prod 60% | Conv 30% | Mon 10%"),
    "Auto Renovação Digital": ("VENDA / RENOVAÇÃO / ASSESSORIA", "Prod 60% | Conv 30% | Mon 10%"),
    "Auto Renovação":         ("VENDA / RENOVAÇÃO / ASSESSORIA", "Prod 60% | Conv 30% | Mon 10%"),
    "Auto Monitorada":        ("VENDA / RENOVAÇÃO / ASSESSORIA", "Prod 60% | Conv 30% | Mon 10%"),
    "Auto Assessoria":        ("VENDA / RENOVAÇÃO / ASSESSORIA", "Prod 60% | Conv 30% | Mon 10%"),
    "Auto Retenção":          ("RETENÇÃO",                       "Ret 60% | Cross 20% | Mon 20%"),
    "Auto Pós-Venda":         ("ENDOSSO (VOZ E DIGITAL)",        "Up 60% | Cross 30% | Mon 10%"),
    "Auto Endosso Digital":   ("ENDOSSO (VOZ E DIGITAL)",        "Up 60% | Cross 30% | Mon 10%"),
    "Auto Endosso Frota":     ("ENDOSSO FROTA",                  "Up 85% | Mon 15%"),
    "Auto Frota":             ("FROTA (VENDA E RENOVAÇÃO)",      "Prod 85% | Mon 15%"),
}

for r_idx, (op, (tipo, pesos)) in enumerate(TIPOS.items(), 2):
    inds = " | ".join(OPERACOES[op])
    fill = PatternFill("solid", fgColor=ROW_FILLS[r_idx % 2])
    for c, val in enumerate([op, tipo, inds, pesos], 1):
        cell = ref.cell(row=r_idx, column=c, value=val)
        cell.fill   = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")

# ─────────────────────────────────────────────
out = "template_metas.xlsx"
wb.save(out)
print(f"Template gerado: {out}  ({len(OPERACOES)} operações, {sum(len(v) for v in OPERACOES.values())} linhas de meta)")
